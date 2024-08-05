import os
from datetime import datetime

from fastapi import UploadFile, HTTPException, status
from openpyxl.worksheet.page import PageMargins

from src.core.database import Order, UnitedOrder
from src.core.settings import BASE_DIR, settings
from src.excel_parser import parse

import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side


def normalize_phone(phone: str):
    reserve_phone = phone
    letters = " +-_()"

    for i in letters:
        phone = phone.replace(i, "")

    if len(phone) == 11:
        phone = "+7" + phone[1::]
        return phone

    elif len(phone) == 10:
        phone = "+7" + phone
        return phone

    return reserve_phone


async def save_file(file: UploadFile):
    content = await file.read()
    date = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
    filename = f"{date}_{file.filename}"
    full_filename = os.path.join(BASE_DIR, 'uploaded_files', filename)

    with open(full_filename, "wb") as f:
        f.write(content)

    return full_filename


async def parse_excel(file: UploadFile, ):

    filename = await save_file(file)

    # json объекты
    try:
        united_orders = parse(filename)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File parsing error: {e}"
        )
    os.remove(filename)
    return united_orders


async def create_payment_list_excel(united_orders: list[UnitedOrder]):
    wb = openpyxl.Workbook()
    wb.create_sheet(title='Первый лист', index=0)

    ns = NamedStyle(name='highlight')
    border = Side(style='thin', color='000000')
    ns.border = Border(left=border, top=border, right=border, bottom=border)

    wb.add_named_style(ns)

    # получаем лист, с которым будем работать
    sheet = wb['Первый лист']
    sheet.page_margins = PageMargins(left=0, right=0, top=0, bottom=0)

    cur_row = 1

    for united_order in united_orders:
        sheet.cell(row=cur_row, column=2, value=united_order.id).style = 'highlight'
        cur_row += 1

        sheet.cell(row=cur_row, column=2, value="ID Заказа").style = 'highlight'
        sheet.cell(row=cur_row, column=3, value="ФИО").style = 'highlight'
        sheet.cell(row=cur_row, column=4, value="Телефон").style = 'highlight'
        sheet.cell(row=cur_row, column=5, value="Дата выдачи").style = 'highlight'
        sheet.cell(row=cur_row, column=6, value="Подпись клиента").style = 'highlight'
        cur_row += 1

        orders = sorted(list(united_order.orders_relationship), key=lambda o: o.customer_name)

        for i in range(len(orders)):

            order = orders[i]

            sheet.cell(row=cur_row, column=1, value=str(i)).style = 'highlight'
            sheet.cell(row=cur_row, column=2, value=order.id).style = 'highlight'
            sheet.cell(row=cur_row, column=3, value=order.customer_name).style = 'highlight'
            sheet.cell(row=cur_row, column=4, value=order.customer_phone).style = 'highlight'
            sheet.cell(row=cur_row, column=5).style = 'highlight'
            sheet.cell(row=cur_row, column=6).style = 'highlight'

            sheet.row_dimensions[cur_row].height = 30

            cur_row += 1

        cur_row += 1


    #sheet.row_dimensions[1].height = 70

    sheet.column_dimensions['A'].width = 2
    sheet.column_dimensions['B'].width = 14
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 14
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 20

    date = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
    filename = f"{date}_payment.xlsx"
    full_filename = os.path.join(BASE_DIR, 'uploaded_files', filename)

    wb.save(full_filename)

    return full_filename



